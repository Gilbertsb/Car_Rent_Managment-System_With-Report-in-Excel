from car_class import *
from openpyxl import *

#I have test my program on picking wrong menu number it can not proceed if you select
#wrong number, it will take you back

all = []     #list to hold class objects
all.append(Car_class("Benz 2xs", "2012", "2014", "RAA454Y", "AVAILABLE", 100, 0, 0))
all.append(Car_class("Audi 9f", "2013", "2015", "RAD34H", "AVAILABLE", 300, 0, 0))
all.append(Car_class("Dodge Ram", "2015", "2017", "RAC354Y", "AVAILABLE", 200, 0, 0))
all.append(Car_class("Dodge Ram", "2015", "2017", "RAC354Y", "AVAILABLE", 200, 0, 0))
all.append(Car_class("Polo cup", "2016", "2018", "RAD56Y", "AVAILABLE", 200, 0, 0))


def add_car():
    car_mod = input("Enter Car model: ")  # helps to add new modle of new car
    year_rlzed = input("Enter a year that Car has been released in: ")  # add year realesed of new car
    year_acqrd = input("Enter a year that Car has been acquired in: ")  # add year acquired of new car
    plate_num = input("Enter a plate number of that Car: ")  # add plate number of new car
    price = eval(input("Enter fixed price/day of this car for rent: "))
    status = 'AVAILABLE'  # assign status to new created key
    prices = int(price)  # assign price to new created key
    money_made = 0  # assign 0 money to new created key
    rent_number = 0  # assign 0 renting times to a new created key

    all.append(Car_class(car_mod, year_rlzed, year_acqrd, plate_num, status, prices, money_made, rent_number))

    wb = load_workbook("Report.xlsx")
    ws = wb.active
    ws['A7'] = car_mod
    ws['B7'] = money_made
    ws['C7'] = rent_number

    wb.save("Report.xlsx")


def rent_car():
    #hold={'x':Car_class(car_model=x,year_released=x, year_acquired=x,plate_number=x,status=x,prices=x,money_made=x,rent_number=x ) for x in all_1}
    index=1
    for c in all: # display of car model and status
        print(index, ':'+ c.car_model,'--',c.status)
        index+=1


        # prompting user to select a car he/she want to rent
    rent = eval(input("Select carn(ex:car1) you want to rent: "))

    car_temp=all[rent-1]

    for i in all: #loop to check availability and allow selection to go smothly
        # checking if car is available
        if car_temp == i and car_temp.status == 'AVAILABLE':
            #asking user to choose b2n fixed/negotiable price
            prc =input('For fixed price/day enter F and for negotiable price enter N : ')
            # checking if it is negotiable price
            if prc.casefold()=='N'.casefold():
                # adding one time to renting times
                i.rent_number=i.rent_number + 1
                # asking user to enter money
                money = input('Enter amount of money you are renting a car: ')
                #adding money
                i.money_made = i.money_made+int(money)
                #changing status to 'ON RENT'
                i.status = 'ON RENT'
                #Dsplaying updates on car
                print('For now ', i.car_model, ' has been rented on ', i.money_made, ' in total')
                print('And it has been rented', i.rent_number, 'times')
                wb = load_workbook("Report.xlsx")
                ws = wb.active
                for row in ws.iter_rows(3):
                    for cell in row:
                        if cell.value == i.car_model:
                            ws.cell(row=cell.row, column=2).value = ws.cell(row=cell.row, column=2).value + int(money)
                            ws.cell(row=cell.row, column=3).value = ws.cell(row=cell.row, column=3).value + 1

                            wb.save("Report.xlsx")

                #chacking if it is fixed prices
            elif prc.casefold()=='F'.casefold():
                # adding one time to renting times
                i.rent_number = i.rent_number + 1
                # adding money
                money_tmp = i.prices+i.money_made
                i.money_made=money_tmp
                # changing status to 'ON RENT'
                i.status = 'ON RENT'
                # Dsplaying updates on car
                print('For now ', i.car_model, ' has been rented on ', i.money_made, ' in total')
                print('And it has been rented', i.rent_number, 'times')

                wb = load_workbook("Report.xlsx")
                ws = wb.active
                for row in ws.iter_rows(3):
                    for cell in row:
                        if cell.value == i.car_model:
                            ws.cell(row=cell.row, column=2).value = ws.cell(row=cell.row, column=2).value + i.prices
                            ws.cell(row=cell.row, column=3).value = ws.cell(row=cell.row, column=3).value + 1

                            wb.save("Report.xlsx")

            #program will display if user doesn't choose any of fixed or negotiable
            else:
                print("You didn't choose any")

               #chacking if user choose a car that is on rent
        elif car_temp == i and car_temp.status=='ON RENT':
            print('This car is on rent now!!!')




def remov_car():
    #text to remind user to selecte a car he/she wants to remove
    print('Select a car you want to remove')

    index = 1
    for v in all:  # display of car model and status
        print(index, ':' + v.car_model)
        index += 1

        #prompting user to select car he/she wants to remove
    remov = eval(input("Select carn(ex:car1) you want to remove: "))
    car_rem = all[remov -1]
    for f in all: #loop to delete cars, dectionsly was conerted into list to delete
        if car_rem == f:  #checking if selected car belongs to dictionaly carmodel
            all.remove(car_rem)
            print("YOUR CAR HAS SUCCESSFULLY REMOVED FROM THE LIST!!")
            index = 1
            for v in all:  # display of car model and status
                print(index, ':' + v.car_model)
                index += 1
def check_trans():
    print("___________________________________")
    print('| Car          Money     Rent times ')  #tabs and decorations
    print("|-----------------------------------")
    for j in all:                           #loop to desplay all cars through car model
        #display car model, money made from them and renting times
        print('|', j.car_model, '--->', j.money_made, '------->', j.rent_number)




    print("-----------------------------------") #decoration
    wb = openpyxl.load_workbook('Report.xlsx')
    type(wb)
#this function display all cars and main details of them
def all_cars():
    #decorations
    print("_______________________________________________________________________")
    print('| Car    |    Plate_number | Released | Acquired | Money | Rent times')
    print("|----------------------------------------------------------------------")
    for op in all: #loops to display all cars through car model
        #display all cars and main details of them
        print('|', op.car_model, '--', op.plate_number, '----', op.year_released, '----',
              op.year_acquired, '------', op.money_made, '-------', op.rent_number)
    print("|----------------------------------------------------------------------")

def put_on_cue():
    index = 1
    for i in all:   #loops to display status
        if i.status =='ON RENT':   #condition to check if car has ON RENT status
            print(index,':',i.car_model,'--',i.status)    #display cars on rent
            index+=1
            #prompting user to select car he/sh wants to put back on cue
    put_back=eval(input('Select carn(ex:car1) you want to put back on cue:'))
    put=all[put_back - 1]
    for x in all: #loop to find car user spacified
        if put == x:   #check if car user spaified is on the list
            x.status = 'AVAILABLE'  #changing status of car
            print("YOUR CAR HAS SUCCESSFULLY ADDED BACK ON CUE!!") #Notifying user
# this function helps to display menu
def main():
    print("USE NUMBER TO SELECT ANY THING FROM MENU")
    print("____________________________________________")
    print("|\t 1.ADD NEW CAR                          |")
    print("|\t 2.RENT A CAR                           |")
    print("|\t 3.REMOVE A CAR                         |")
    print("|\t 4.CHECK NUMBER OF TIMES A CAR HAS      |")
    print("|\t   BEEN RENTED, AND MONEY MADE FROM IT  |")
    print("|\t 5.DISPLAY ALL CARS IN YOUR STORE       |")
    print("|\t 6.PUT A CAR BACK ON CUE                |")
    print("|___________________________________________|")
     #prompting user to select one of the menu
    selection = input("Select menu with using number(1-6): ")

    if selection == '1': #checking if user selected one
        add_car()        #calling function to add car
    elif selection == '2':  #checking if user selected two
        rent_car()          #calling function to rent car
    elif selection == '3':  #checking if user selected three
        remov_car()         #calling function to remove car
    elif selection == '4':  #checking if user selected four
        check_trans()       #calling function to check cars, money made and renting times
    elif selection == '5':  #checking if user selected five
        all_cars()          #calling function to display cars
    elif selection == '6':  #checking if user selected six
        put_on_cue()        #calling function to put car back on cue
    else:
        #if user doesn't  choose any of above numbers
        print("NO CHOICE WE HAVE THAT LOOK LIKE THAT")
def exit_this():
    exit("THANK YOU FOR USING THIS SOFTWARE SEE YOUUU!!!")





